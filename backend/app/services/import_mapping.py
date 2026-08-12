"""Header detection and column mapping for the smart import wizard.

The generic import engine (``app.api.v1.dataio``) originally required a file whose
header row matched our template *exactly*. Every organisation names its columns
differently ("Risk Desc", "Probability", "Existing Controls"), so onboarding meant the
client hand-rewriting their spreadsheet. This module removes that step: it reads the
file the client already has, works out which of their columns feed which of our fields,
and hands the caller a mapping it can show, correct and save.

Everything here is pure — no DB, no FastAPI — so the matching rules are unit-testable
and reusable from a future AI-assisted mapper.

Pipeline
--------
1. :func:`load_table` turns raw CSV text *or* an XLSX workbook into a rectangular grid,
   skipping any banner rows above the real header (banks routinely export a title row).
2. :func:`suggest_mapping` scores every (their column -> our column) pair through five
   tiers of increasing looseness and greedily assigns the best non-conflicting set.
3. :func:`apply_mapping` rewrites one of their rows into a row the existing import
   engine already understands, so nothing downstream changes.
"""
from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass
from difflib import SequenceMatcher
from typing import TYPE_CHECKING

if TYPE_CHECKING:  # pragma: no cover - type hints only
    from app.services.import_registry import Column

__all__ = [
    "MappingSuggestion",
    "TableData",
    "apply_mapping",
    "custom_field_model_key",
    "load_table",
    "normalise",
    "suggest_mapping",
]

# How many leading rows we are willing to skip while hunting for the header.
_MAX_BANNER_ROWS = 10
# A header row must carry at least this many filled cells (guards against a title banner).
_MIN_HEADER_CELLS = 2

# Words that carry no signal when comparing column names.
_STOPWORDS = frozenset({"the", "of", "a", "an", "and", "or", "for", "to", "in", "on", "no"})

# Confidence floors for the looser tiers. Anything below `_MIN_CONFIDENCE` is reported
# as unmapped rather than guessed at — a wrong silent mapping is worse than none.
_MIN_CONFIDENCE = 0.50
_FUZZY_FLOOR = 0.82  # difflib ratio below this is noise


# ---------------------------------------------------------------------------
# Synonym dictionary — keyed by OUR field name, values are normalised phrasings
# seen in real client spreadsheets. Order does not matter; matching is set-based.
# ---------------------------------------------------------------------------
SYNONYMS: dict[str, tuple[str, ...]] = {
    # --- universal identity / description ---------------------------------
    "title": ("name", "risk name", "risk title", "control name", "issue title", "subject", "particulars"),
    "name": ("title", "risk name", "asset name", "control name", "vendor name", "system name", "particulars"),
    "description": ("desc", "details", "detail", "narrative", "risk description", "explanation", "remarks", "summary"),
    "reference": ("ref", "ref no", "id", "code", "serial", "sr no", "s no", "sno", "risk id", "control id", "unique id"),
    "category": ("type", "risk category", "risk type", "classification", "class", "domain", "group", "area"),
    # --- ownership ---------------------------------------------------------
    "owner": ("risk owner", "owner name", "responsible", "responsible person", "accountable", "custodian", "assigned to"),
    "treatment_owner": ("action owner", "mitigation owner", "remediation owner", "responsible for action"),
    "action_owner": ("owner of action", "responsible for action", "assigned to", "remediation owner"),
    "lead_auditor": ("auditor", "audit lead", "lead", "engagement lead"),
    "information_owner": ("data owner", "business owner", "asset owner"),
    "business_unit": ("department", "dept", "bu", "function", "division", "unit", "branch"),
    # --- qualitative scoring ----------------------------------------------
    "inherent_likelihood": (
        "likelihood", "probability", "prob", "chance", "frequency of occurrence",
        "inherent probability", "gross likelihood", "likelihood before controls",
    ),
    "inherent_impact": (
        "impact", "consequence", "severity", "magnitude", "inherent consequence",
        "gross impact", "impact before controls",
    ),
    "residual_likelihood": (
        "post control likelihood", "likelihood after controls", "net likelihood",
        "residual probability", "controlled likelihood",
    ),
    "residual_impact": (
        "post control impact", "impact after controls", "net impact",
        "residual consequence", "controlled impact",
    ),
    "severity": ("rating", "priority", "criticality", "risk rating", "grade", "level"),
    "criticality": ("critical", "criticality level", "importance", "business criticality", "priority"),
    # --- treatment / remediation ------------------------------------------
    "treatment_strategy": ("treatment", "response", "risk response", "strategy", "mitigation strategy", "action type"),
    "treatment_description": (
        "mitigation plan", "action plan", "treatment plan", "remediation plan",
        "mitigation", "remediation", "corrective action", "management action",
    ),
    "treatment_deadline": ("target date", "deadline", "completion date", "target completion", "due"),
    "treatment_cost": ("cost", "budget", "estimated cost", "mitigation cost"),
    "recommendation": ("auditor recommendation", "suggested action", "advice"),
    "management_response": ("management comment", "auditee response", "response from management"),
    # --- dates -------------------------------------------------------------
    "due_date": ("due", "target date", "deadline", "completion date", "expected closure"),
    "next_review_date": ("review date", "next review", "review due", "next assessment"),
    "last_review_date": ("last reviewed", "previous review", "date of last review"),
    "review_frequency": ("frequency", "review cycle", "review period", "periodicity", "cycle"),
    "assessment_date": ("date of assessment", "assessed on", "date assessed"),
    "report_date": ("date of report", "reported on", "issue date"),
    "period_start": ("from date", "start", "start date", "audit period from"),
    "period_end": ("to date", "end", "end date", "audit period to"),
    "closed_date": ("date closed", "closure date", "resolved on"),
    # --- status ------------------------------------------------------------
    "status": ("risk status", "state", "current status", "stage", "progress"),
    "workflow_status": ("approval status", "review status", "workflow state"),
    # --- quantitative ------------------------------------------------------
    "annual_loss_frequency": ("loss frequency", "events per year", "alf", "lef", "loss event frequency"),
    "single_loss_expectancy": ("sle", "loss per event", "single loss", "expected loss per event"),
    "currency": ("ccy", "currency code"),
    # --- continuity / BIA --------------------------------------------------
    "rto_hours": ("rto", "recovery time objective", "recovery time"),
    "rpo_hours": ("rpo", "recovery point objective", "recovery point"),
    "mtpd_hours": ("mtpd", "maximum tolerable period of disruption", "max outage"),
    # --- assets ------------------------------------------------------------
    "hostname": ("host", "server name", "machine name", "device name"),
    "ip_address": ("ip", "ip addr", "address"),
    "serial_number": ("serial", "sn", "asset serial"),
    "manufacturer": ("vendor", "make", "oem", "brand"),
    "model_number": ("model", "model no"),
    "os_version": ("os", "operating system", "platform"),
    "location": ("site", "premises", "data centre", "data center", "city"),
    "environment": ("env", "tier", "deployment"),
    "replacement_cost": ("cost", "value", "asset value", "purchase cost", "book value"),
    # --- links (multi-token reference cells) -------------------------------
    "assets": ("asset", "affected asset", "asset name", "system", "application", "systems affected"),
    "controls": (
        "control", "existing control", "existing controls", "mitigating control",
        "mitigating controls", "control name", "controls in place", "current controls",
    ),
    "threats": ("threat", "threat name", "threat source", "threat actor"),
    "vulnerabilities": ("vulnerability", "weakness", "vuln", "gap"),
    "policies": ("policy", "related policy", "governing policy"),
    "incidents": ("incident", "related incident", "event"),
    "risks": ("risk", "related risk", "linked risk", "risk title"),
    "requirements": ("requirement", "clause", "control objective", "reference control"),
    "process": ("business process", "process name", "activity"),
    "vendor": ("supplier", "third party", "service provider", "vendor name", "outsourcing partner"),
    # --- contact -----------------------------------------------------------
    "email": ("email address", "e mail", "mail"),
    "phone": ("phone number", "contact number", "mobile", "telephone", "contact"),
    "contact_name": ("contact person", "focal person", "poc", "point of contact"),
}

# Nouns that appear in nearly every column of a given register and therefore carry no
# discriminating signal there ("Risk Title", "Risk Category", "Risk Owner" in a risk sheet).
_RESOURCE_NOUNS: dict[str, tuple[str, ...]] = {
    "risks": ("risk",),
    "controls": ("control",),
    "it-assets": ("asset", "it"),
    "information-assets": ("asset", "information"),
    "vendors": ("vendor", "supplier"),
    "policies": ("policy",),
    "incidents": ("incident",),
    "issues": ("issue",),
    "audit-engagements": ("audit", "engagement"),
    "audit-findings": ("audit", "finding"),
    "threats": ("threat",),
    "vulnerabilities": ("vulnerability",),
    "evidence": ("evidence",),
    "projects": ("project",),
    "goals": ("goal",),
    "requirements": ("requirement",),
    "loss-events": ("loss", "event"),
    "kris": ("kri", "indicator"),
    "bia-assessments": ("bia", "assessment"),
    "continuity-plans": ("plan", "continuity"),
}


# ---------------------------------------------------------------------------
# Normalisation
# ---------------------------------------------------------------------------
_LEADING_NUMBER = re.compile(r"^\s*\d+\s*[.)\-]\s*")
_PARENTHETICAL = re.compile(r"\([^)]*\)")
_NON_ALNUM = re.compile(r"[^a-z0-9]+")


def normalise(value: str) -> str:
    """Reduce a column name to comparable form.

    Strips a leading list number ("1. Risk Title"), drops parenthetical hints
    ("Impact (1-5)"), folds punctuation/underscores to spaces and lowercases. This is
    what makes ``Risk_Description``, ``risk description`` and ``2) Risk Description``
    all compare equal.
    """
    text = _LEADING_NUMBER.sub("", str(value or "").strip().lower())
    text = _PARENTHETICAL.sub(" ", text)
    text = _NON_ALNUM.sub(" ", text)
    return re.sub(r"\s+", " ", text).strip()


def _tokens(value: str, *, drop: frozenset[str] = frozenset()) -> frozenset[str]:
    """Meaningful word set of a normalised column name."""
    return frozenset(
        word for word in normalise(value).split() if word not in _STOPWORDS and word not in drop
    )


def _jaccard(left: frozenset[str], right: frozenset[str]) -> float:
    if not left or not right:
        return 0.0
    return len(left & right) / len(left | right)


# ---------------------------------------------------------------------------
# Table loading (CSV text or XLSX bytes -> a clean grid)
# ---------------------------------------------------------------------------
@dataclass(frozen=True)
class TableData:
    """A loaded spreadsheet, already trimmed to its real header row.

    ``csv`` is re-emitted canonical CSV text — every later step (preview, import)
    consumes this, so XLSX and banner rows are handled exactly once, here.
    """

    headers: list[str]
    rows: list[list[str]]
    csv: str
    header_row_index: int  # 0-based index of the header row in the ORIGINAL file
    sheet_names: list[str]
    sheet: str

    @property
    def row_count(self) -> int:
        return len(self.rows)


def _detect_header_row(grid: list[list[str]]) -> int:
    """Index of the first row that behaves like a header.

    Client exports frequently open with a title banner ("Annexure-A: Risk Register")
    or a blank line. A header row is the first row within the first few that carries
    at least two filled cells and is followed by at least one more non-empty row.
    """
    limit = min(len(grid), _MAX_BANNER_ROWS)
    for index in range(limit):
        filled = sum(1 for cell in grid[index] if str(cell).strip())
        if filled < _MIN_HEADER_CELLS:
            continue
        has_data_below = any(
            any(str(cell).strip() for cell in row) for row in grid[index + 1 :]
        )
        if has_data_below:
            return index
    return 0


def _dedupe_headers(headers: list[str]) -> list[str]:
    """Make blank/duplicate headers unique so a dict-per-row stays lossless."""
    seen: dict[str, int] = {}
    out: list[str] = []
    for position, raw in enumerate(headers):
        header = str(raw or "").strip() or f"Column {position + 1}"
        if header in seen:
            seen[header] += 1
            header = f"{header} ({seen[header]})"
        else:
            seen[header] = 1
        out.append(header)
    return out


def _grid_to_table(
    grid: list[list[str]], *, sheet_names: list[str], sheet: str
) -> TableData:
    if not grid:
        return TableData([], [], "", 0, sheet_names, sheet)

    header_index = _detect_header_row(grid)
    headers = _dedupe_headers(grid[header_index])
    width = len(headers)

    rows: list[list[str]] = []
    for raw_row in grid[header_index + 1 :]:
        if not any(str(cell).strip() for cell in raw_row):
            continue  # skip blank separator rows
        row = [str(cell).strip() if cell is not None else "" for cell in raw_row[:width]]
        row.extend([""] * (width - len(row)))
        rows.append(row)

    buf = io.StringIO()
    writer = csv.writer(buf, lineterminator="\n")
    writer.writerow(headers)
    writer.writerows(rows)
    return TableData(headers, rows, buf.getvalue(), header_index, sheet_names, sheet)


def load_csv(content: str) -> TableData:
    grid = [list(row) for row in csv.reader(io.StringIO(content))]
    return _grid_to_table(grid, sheet_names=[], sheet="")


def _xl_cell(value: object) -> str:
    """Render one Excel cell as text, keeping whole numbers integral.

    Excel stores every number as a float, so a likelihood of 3 arrives as ``3.0`` and
    would fail integer coercion downstream. Dates arrive as datetimes and must reach
    the importer as ISO strings.
    """
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if hasattr(value, "isoformat"):
        text = value.isoformat()
        # A date-only value round-trips as YYYY-MM-DDT00:00:00; trim to the date.
        return text[:10] if text.endswith("T00:00:00") else text
    return str(value).strip()


def load_xlsx(data: bytes, sheet: str | None = None) -> TableData:
    """Read an .xlsx workbook into a table. Raises ``ValueError`` with a clean message."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - depends on the deployed env
        raise ValueError(
            "Excel import needs the 'openpyxl' package. Save the file as CSV, "
            "or install openpyxl on the server."
        ) from exc

    try:
        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 - any openpyxl failure is a bad upload
        raise ValueError(f"Could not read the Excel file: {exc}") from exc

    try:
        sheet_names = list(workbook.sheetnames)
        if not sheet_names:
            raise ValueError("The workbook contains no sheets")
        chosen = sheet if sheet in sheet_names else sheet_names[0]
        worksheet = workbook[chosen]
        grid = [[_xl_cell(cell) for cell in row] for row in worksheet.iter_rows(values_only=True)]
    finally:
        workbook.close()

    return _grid_to_table(grid, sheet_names=sheet_names, sheet=chosen)


def load_table(
    *, content: str | None = None, data: bytes | None = None, sheet: str | None = None
) -> TableData:
    """Load either raw CSV text (``content``) or XLSX bytes (``data``)."""
    if data is not None:
        return load_xlsx(data, sheet)
    return load_csv(content or "")


# ---------------------------------------------------------------------------
# Mapping suggestion
# ---------------------------------------------------------------------------
@dataclass(frozen=True)
class MappingSuggestion:
    """One proposed ``their column -> our column`` pair."""

    source: str  # header as it appears in the client's file
    target: str  # our canonical Column.header
    field: str  # the Create-schema field it feeds
    confidence: float  # 0.0 - 1.0
    reason: str  # short, human-readable justification shown in the wizard

    @property
    def band(self) -> str:
        """UI grouping: ``high`` maps silently, ``medium``/``low`` ask for a look."""
        if self.confidence >= 0.90:
            return "high"
        if self.confidence >= 0.70:
            return "medium"
        return "low"


def _score(source: str, column: "Column", drop: frozenset[str]) -> tuple[float, str] | None:
    """Best (confidence, reason) for one source header against one of our columns.

    Five tiers, tried in order of decreasing certainty. Returns ``None`` when even the
    loosest tier finds nothing — the column is then reported unmapped, which the user
    resolves explicitly. Guessing here would silently load data into the wrong field.
    """
    raw = str(source or "").strip()
    if not raw:
        return None

    target_header = column.header
    target_field = column.field
    if raw == target_header:
        return 1.0, "exact match"

    source_norm = normalise(raw)
    header_norm = normalise(target_header)
    field_norm = normalise(target_field)
    if not source_norm:
        return None

    # Tier 2 — same name once punctuation and case are ignored.
    if source_norm in (header_norm, field_norm):
        return 0.97, "matches after normalising case and punctuation"

    # Tier 3 — a known phrasing for this field. Link columns carry a friendly header
    # ("controls") over an ids-suffixed field ("control_ids"), so both are looked up.
    synonyms = tuple(
        dict.fromkeys(SYNONYMS.get(target_field, ()) + SYNONYMS.get(target_header, ()))
    )
    if source_norm in synonyms:
        return 0.90, f"known alternative name for '{target_header}'"

    # Tier 3b — the column's own name appears inside a longer client header, e.g.
    # "Inherent Likelihood Score" -> inherent_likelihood. Uses the header rather than the
    # field so a link column is not judged on its meaningless "_ids" suffix.
    identity_tokens = _tokens(target_header) or _tokens(target_field)
    source_tokens = _tokens(source_norm)
    if identity_tokens and identity_tokens <= source_tokens:
        return 0.85, f"contains every word of '{target_header}'"

    # Tier 4 — token overlap once the register's own noun is discounted, so
    # "Risk Category" and "Category" agree in a risk import.
    lean_source = _tokens(source_norm, drop=drop)
    lean_target = _tokens(target_header, drop=drop) | _tokens(target_field, drop=drop)
    if lean_source and lean_target:
        if lean_source == lean_target:
            return 0.88, "same words in a different order"
        overlap = _jaccard(lean_source, lean_target)
        if overlap >= 0.5:
            return 0.50 + 0.30 * overlap, "shares most of its words"

    # Tier 4b — a synonym phrase overlapping on tokens (e.g. "Existing Controls (list)").
    for phrase in synonyms:
        phrase_tokens = _tokens(phrase, drop=drop)
        if phrase_tokens and phrase_tokens <= source_tokens:
            return 0.80, f"reads like '{phrase}'"

    # Tier 5 — character-level similarity, for typos and abbreviations.
    ratio = max(
        SequenceMatcher(None, source_norm, header_norm).ratio(),
        SequenceMatcher(None, source_norm, field_norm).ratio(),
    )
    if ratio >= _FUZZY_FLOOR:
        return ratio * 0.75, "spelled almost the same"

    return None


def suggest_mapping(
    source_headers: list[str], columns: list["Column"], *, resource: str = ""
) -> tuple[list[MappingSuggestion], list[str], list[str]]:
    """Propose a mapping from the client's headers onto our columns.

    Returns ``(suggestions, unmapped_source_headers, unfilled_target_headers)``.
    Assignment is greedy on confidence and strictly one-to-one in both directions: one
    of their columns can never feed two of our fields, and two of their columns can
    never fight over one of ours. Ties break on the file's own column order, so the
    result is deterministic for a given file.
    """
    drop = frozenset(_RESOURCE_NOUNS.get(resource, ()))

    candidates: list[tuple[float, int, int, str]] = []
    for source_index, source in enumerate(source_headers):
        for column_index, column in enumerate(columns):
            scored = _score(source, column, drop)
            if scored is None:
                continue
            confidence, reason = scored
            if confidence < _MIN_CONFIDENCE:
                continue
            candidates.append((confidence, source_index, column_index, reason))

    # Highest confidence first; stable on the file's column order for reproducibility.
    candidates.sort(key=lambda item: (-item[0], item[1], item[2]))

    taken_sources: set[int] = set()
    taken_targets: set[int] = set()
    suggestions: list[MappingSuggestion] = []
    for confidence, source_index, column_index, reason in candidates:
        if source_index in taken_sources or column_index in taken_targets:
            continue
        taken_sources.add(source_index)
        taken_targets.add(column_index)
        column = columns[column_index]
        suggestions.append(
            MappingSuggestion(
                source=source_headers[source_index],
                target=column.header,
                field=column.field,
                confidence=round(confidence, 3),
                reason=reason,
            )
        )

    suggestions.sort(key=lambda s: source_headers.index(s.source))
    unmapped = [h for i, h in enumerate(source_headers) if i not in taken_sources]
    unfilled = [c.header for i, c in enumerate(columns) if i not in taken_targets]
    return suggestions, unmapped, unfilled


# ---------------------------------------------------------------------------
# Applying a mapping
# ---------------------------------------------------------------------------
def apply_mapping(row: dict[str, str | None], mapping: dict[str, str]) -> dict[str, str | None]:
    """Rewrite one of the client's rows into our canonical header space.

    ``mapping`` is ``{their header: our header}``. Columns absent from the mapping are
    dropped, which is what makes an unmapped column explicitly ignored rather than
    accidentally matched by name further down the pipeline.
    """
    out: dict[str, str | None] = {}
    for source, target in mapping.items():
        if not target:
            continue
        out[target] = row.get(source)
    return out


# ---------------------------------------------------------------------------
# Custom-field model key
# ---------------------------------------------------------------------------
_CAMEL_BOUNDARY = re.compile(r"(?<!^)(?=[A-Z])")

# Models whose custom-field key predates this derivation and does not match it. Getting
# one of these wrong would reject custom-field mapping for a register that in fact
# supports it, so the exceptions are listed rather than inferred.
_MODEL_KEY_OVERRIDES: dict[str, str] = {
    "ExceptionRecord": "exception",
}


def custom_field_model_key(model: type) -> str:
    """The ``CUSTOM_FIELD_MODELS`` key for a SQLAlchemy model class.

    Custom fields are addressed by a snake_case model name (``risk``,
    ``audit_engagement``) while the import registry holds the class itself. The two
    agree by convention for all but the handful of names in ``_MODEL_KEY_OVERRIDES``,
    so a resource never has to declare the key twice.
    """
    name = model.__name__
    if name in _MODEL_KEY_OVERRIDES:
        return _MODEL_KEY_OVERRIDES[name]
    return _CAMEL_BOUNDARY.sub("_", name).lower()
