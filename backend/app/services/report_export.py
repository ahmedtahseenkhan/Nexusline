"""Tabular exports for the report builder: Excel and CSV.

The PDF is the committee pack; these are for the analyst who wants to pivot. Two
things distinguish them from a raw dump:

* **The Excel carries a Parameters sheet** — subject, every filter that was set, when
  it was run and by whom, and the summary counts. A spreadsheet that circulates without
  its parameters becomes "the risk register" in someone's inbox; with them, it is
  provably "critical risks in Digital Banking as at 5 September".
* **Nothing is interpreted.** Cells are the same strings the on-screen table shows, so
  a number that reads 25 on screen is 25 in the file, not a re-derived value that may
  differ by a rounding rule.

``openpyxl`` is imported lazily, as ``reportlab`` is in the PDF module, so a server
without it still serves CSV and gives a clear message rather than failing to boot.
"""
from __future__ import annotations

import io
from datetime import datetime, timezone

from fastapi import HTTPException, status

from app.services import csv_io


def to_csv(headers: list[str], rows: list[list]) -> str:
    return csv_io.export_csv([dict(zip(headers, r)) for r in rows], headers)


def to_xlsx(
    *,
    title: str,
    org_name: str,
    subject_label: str,
    run_by: str,
    params: list[tuple[str, str]],
    summary: dict[str, dict[str, int]],
    headers: list[str],
    rows: list[list],
) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ModuleNotFoundError as exc:  # pragma: no cover
        raise HTTPException(
            status_code=status.HTTP_501_NOT_IMPLEMENTED,
            detail="Excel export requires the 'openpyxl' package on the server; CSV is available.",
        ) from exc

    wb = Workbook()

    # ----------------------------------------------------------------- report
    ws = wb.active
    ws.title = "Report"
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="1D4FD7")
    ws.append(headers)
    for cell in ws[1]:
        cell.font = head_font
        cell.fill = head_fill
        cell.alignment = Alignment(vertical="center")
    for row in rows:
        ws.append([_cell(v) for v in row])
    ws.freeze_panes = "A2"
    if rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"
    # Width from content, capped so a description column does not become a screen.
    for index, header in enumerate(headers, start=1):
        longest = max([len(str(header))] + [len(str(_cell(r[index - 1]))) for r in rows[:500]])
        ws.column_dimensions[get_column_letter(index)].width = min(max(10, longest + 2), 60)

    # ------------------------------------------------------------- parameters
    ps = wb.create_sheet("Parameters")
    bold = Font(bold=True)
    generated = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    for label, value in (
        ("Report", title),
        ("Organisation", org_name),
        ("Subject", subject_label),
        ("Generated", generated),
        ("Run by", run_by),
        ("Rows", str(len(rows))),
    ):
        ps.append([label, value])
        ps.cell(row=ps.max_row, column=1).font = bold
    ps.append([])
    ps.append(["Filters applied", "" if params else "None — whole register"])
    ps.cell(row=ps.max_row, column=1).font = bold
    for label, value in params:
        ps.append([f"  {label}", value])
    for section, counts in summary.items():
        ps.append([])
        ps.append([section, ""])
        ps.cell(row=ps.max_row, column=1).font = bold
        for key, n in counts.items():
            ps.append([f"  {key}", n])
    ps.column_dimensions["A"].width = 28
    ps.column_dimensions["B"].width = 60

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _cell(value):
    """Keep numbers numeric so Excel can sum them; everything else as shown on screen."""
    if value is None:
        return ""
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return value
    return str(value)
